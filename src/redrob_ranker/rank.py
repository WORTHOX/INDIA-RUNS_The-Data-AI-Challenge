"""Turn precomputed feature artifacts into final top-100 output files.

Ranking is intentionally fast and deterministic: verify the candidate/JD hashes,
load the parquet artifact, calculate final scores, sort, explain, and write the
submission files.
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from redrob_ranker.job_description import compute_job_text_hash, read_job_description_text
from redrob_ranker.reasoning.evidence_builder import build_reasoning
from redrob_ranker.scoring.fit_score import compute_fit_score
from redrob_ranker.scoring.semantic import compute_semantic_support
from redrob_ranker.scoring.suspiciousness import compute_suspiciousness_penalty
from redrob_ranker.utils.hash_guard import read_dataset_hash, verify_dataset_hash
from redrob_ranker.utils.io import read_feature_frame


def _score_row(row: pd.Series) -> dict:
    """Compute final rankable score components for one feature row."""

    fit_score = compute_fit_score(row)
    semantic_support = compute_semantic_support(row)
    suspiciousness_penalty = compute_suspiciousness_penalty(
        row,
        title_features=row,
        skill_features=row,
        honeypot_features=row,
    )
    # Final score is normalized for judge readability. The raw formula can
    # exceed 1.0 when a strong candidate also has excellent availability.
    final_score = min(
        1.0,
        max(
            0.0,
            (fit_score + semantic_support) * float(row.get("availability_multiplier", 1.0))
            - suspiciousness_penalty,
        ),
    )
    return {
        "fit_score": round(fit_score, 6),
        "semantic_support": round(semantic_support, 6),
        "suspiciousness_penalty": round(suspiciousness_penalty, 6),
        "score": round(final_score, 6),
    }


def score_feature_frame(feature_frame: pd.DataFrame) -> pd.DataFrame:
    scored_rows = []
    for _, row in feature_frame.iterrows():
        scores = _score_row(row)
        enriched_row = row.to_dict()
        enriched_row.update(scores)
        scored_rows.append(enriched_row)
    return pd.DataFrame(scored_rows)


def build_ranked_submission(feature_frame: pd.DataFrame, *, top_n: int = 100) -> pd.DataFrame:
    """Score all candidates, sort deterministically, and return submission columns."""

    scored_frame = score_feature_frame(feature_frame)
    scored_frame = scored_frame.sort_values(by=["score", "candidate_id"], ascending=[False, True]).head(top_n).copy()
    scored_frame["rank"] = range(1, len(scored_frame) + 1)
    scored_frame["reasoning"] = [
        build_reasoning(record, rank)
        for rank, record in zip(scored_frame["rank"], scored_frame.to_dict("records"), strict=False)
    ]
    return scored_frame[["candidate_id", "rank", "score", "reasoning"]]


def write_submission_csv(submission_frame: pd.DataFrame, output_path: str | Path) -> None:
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    submission_frame.to_csv(destination, index=False, encoding="utf-8")


def default_xlsx_output_path(csv_output_path: str | Path) -> Path:
    """Return the sidecar XLSX path for a CSV submission path."""

    return Path(csv_output_path).with_suffix(".xlsx")


def write_submission_xlsx(submission_frame: pd.DataFrame, output_path: str | Path) -> None:
    """Write the ranked submission table as a readable Excel workbook."""

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        submission_frame.to_excel(writer, index=False, sheet_name="Ranked Candidates")
        worksheet = writer.sheets["Ranked Candidates"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        column_widths = {
            "A": 18,  # candidate_id
            "B": 10,  # rank
            "C": 12,  # score
            "D": 120,  # reasoning
        }
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for cell in worksheet["C"][1:]:
            cell.number_format = "0.000000"
        for row in worksheet.iter_rows(min_row=2, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=cell.column == 4, vertical="top")


def write_submission_outputs(
    submission_frame: pd.DataFrame,
    csv_output_path: str | Path,
    *,
    xlsx_output_path: str | Path | None = None,
) -> tuple[Path, Path]:
    """Write both official CSV and reviewer-friendly XLSX outputs."""

    csv_path = Path(csv_output_path)
    xlsx_path = Path(xlsx_output_path) if xlsx_output_path else default_xlsx_output_path(csv_path)
    write_submission_csv(submission_frame, csv_path)
    write_submission_xlsx(submission_frame, xlsx_path)
    return csv_path, xlsx_path


def run_ranking(
    candidates_path: str | Path,
    artifacts_dir: str | Path,
    output_path: str | Path,
    *,
    job_description_path: str | Path | None = None,
    xlsx_output_path: str | Path | None = None,
    top_n: int = 100,
) -> pd.DataFrame:
    """Validate artifacts, rank candidates, write outputs, and return the submission frame."""

    artifacts_path = Path(artifacts_dir)
    feature_path = artifacts_path / "candidate_features.parquet"
    hash_path = artifacts_path / "dataset.hash"
    job_hash_path = artifacts_path / "job.hash"

    # These guards are intentionally before reading the parquet artifact. They
    # prevent using stale features with a changed candidate file or JD.
    verify_dataset_hash(candidates_path, hash_path)
    if job_hash_path.exists():
        job_description_text = read_job_description_text(job_description_path)
        expected_hash = read_dataset_hash(job_hash_path)
        actual_hash = compute_job_text_hash(job_description_text)
        if actual_hash != expected_hash:
            raise ValueError(f"Text hash mismatch: expected {expected_hash}, found {actual_hash}")
    feature_frame = read_feature_frame(feature_path)
    submission_frame = build_ranked_submission(feature_frame, top_n=top_n)
    write_submission_outputs(submission_frame, output_path, xlsx_output_path=xlsx_output_path)
    return submission_frame


def main() -> None:
    parser = argparse.ArgumentParser(description="Rank candidates into Redrob submission CSV and XLSX files.")
    parser.add_argument("--candidates", required=True, help="Path to candidates.jsonl")
    parser.add_argument(
        "--job-description",
        required=False,
        help="Path to the JD file used for precompute. Supports .txt/.md and .docx.",
    )
    parser.add_argument("--artifacts", required=True, help="Directory containing precomputed artifacts")
    parser.add_argument("--participant-id", required=True, help="Participant or team identifier")
    parser.add_argument("--out", required=False, help="Optional output csv path")
    parser.add_argument(
        "--xlsx-out",
        required=False,
        help="Optional output xlsx path. Defaults to the CSV output path with .xlsx suffix.",
    )
    args = parser.parse_args()

    output_path = args.out or f"{args.participant_id}.csv"
    run_ranking(
        args.candidates,
        args.artifacts,
        output_path,
        job_description_path=args.job_description,
        xlsx_output_path=args.xlsx_out,
    )


if __name__ == "__main__":
    main()
